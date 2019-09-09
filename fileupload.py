import win32ui
import win32con
import win32gui
import pandas as pd
import getpass
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import openpyxl
import time
from tkinter import Tk
from tkinter.filedialog import askopenfilename
from tkinter import messagebox
import pyodbc
import os
from tkinter import *

#val=""
def Open_window(r_file):
    while True:
        try:
            w=win32gui.FindWindow('#32770','Open')
            w1=win32gui.FindWindowEx(w,0, "ComboBoxEx32",None)
            w2=win32gui.FindWindowEx(w1,0, "ComboBox",None)
            w3=win32gui.FindWindowEx(w2,0, "Edit",None)
            w4=win32ui.CreateWindowFromHandle(w3)
            w4.SendMessage(win32con.WM_SETTEXT,0, r_file)
            b=win32gui.FindWindowEx(w,0, "Button",'&Open')
            b1=win32ui.CreateWindowFromHandle(b)
            b1.SendMessage(win32con.BM_CLICK,0,0)
            break
        except:
            continue
def Test_Agile(driver):
    driver.get("https://agileplm-test.flex.com/Agile")

def Prod_Agile(driver):
    driver.get("https://agileplm.flex.com/Agile")

def User_pass(loc,driver):
    username = getpass.getuser()
    da = pd.ExcelFile(loc)
    data=pd.read_excel(da,'CONFIG')
    newdata=data.dropna()
    ldata=newdata[newdata.columns[1]].tolist()
    driver.get(ldata[3])
    user = driver.find_element_by_id("j_username")
    user.send_keys(ldata[0])
    Pass = driver.find_element_by_id("j_password")
    Pass.send_keys(ldata[1])
    Pass.send_keys(Keys.RETURN)
    return ldata[0],ldata[2]

def tj():
    def getDate():
        global val
        val = E3.get() +"-"+E2.get()+"-"+E1.get()
        root.quit()
        root.destroy()
    root = Tk()
    root.title('Licence Manager')
    label1 = Label( root, text="-",width=20) 
    E1 = Entry(root, bd =5)
    E1.place(width=150,height=50)
    label2 = Label( root, text="-")
    E2 = Entry(root, bd =5)
    E2.place(width=150,height=50)
    label3 = Label( root, text="-")
    E3 = Entry(root, bd =5)
    E3.place(width=150,height=50)
    submit = Button(root, text ="Submit", command = getDate ,width=20,height=2)
    E1.pack(side =RIGHT)
    label2.pack(side =RIGHT)
    E2.pack(side =RIGHT)
    label3.pack(side =RIGHT)
    E3.pack(side =RIGHT)
    submit.pack()
    root.mainloop()

Tk().withdraw()
c_name = os.environ['COMPUTERNAME']
cnxn = pyodbc.connect("DRIVER={SQL Server};"
                          "Server=gssnt022.asia.ad.flextronics.com;"
                          "Database=GBS_EQ;"
                          "uid=EQ;"
                          "pwd=GBSEQ123;")
cursor = cnxn.cursor()
cursor.execute("SELECT * From GBS_EQ_APPLIC_MGMT WHERE Application_Name ='AGILE_DOC_A/R' AND  System_Name = ?",(c_name))
row = cursor.fetchone()
if row is None:
    messagebox.showinfo("Alert","System is not Registred. Please enter the License Key")
    tj()
    cursor.execute("SELECT * From GBS_EQ_APPLIC_MGMT WHERE Application_Name = 'AGILE_DOC_A/R' AND Licence_Key=? AND System_Name =?", (val,""))
    row1=cursor.fetchone()
    if row1 is None:
        messagebox.showerror("Error", "You have entered wrong License Key!!!")
        exit()
    else:
        cursor.execute("UPDATE GBS_EQ_APPLIC_MGMT SET System_Name=? , Status='Active' WHERE Licence_Key=?", (c_name,val))
        cnxn.commit()
cursor.execute("SELECT * FROM VERSION_CONTROL WHERE Tool=?",
                   ("AGILE_DOC_A/R"))
row = cursor.fetchone()
if row is not None:
    Con_ver=row[1]
    cnxn.close()
else:
    exit()
if Con_ver!="V1.0":
    exit()
messagebox.showinfo("Input","Please select the Input File")
filename = askopenfilename()
loc = filename
#path=r"C:\Users\gssharik\Downloads\file\Item"
#loc=r"C:\Users\gssharik\Downloads\file\Book10.xlsx"
inp=openpyxl.load_workbook(loc)
wks = inp['MAIN']
n=wks.max_row+1
driver = webdriver.Chrome(r"\\gssnt022.asia.ad.flextronics.com\GBS-EQ\Applications\Local Apps\RPA\chromedriver.exe")
try:
    userid,path=User_pass(loc,driver)
    window_before = driver.window_handles[0]
    driver.close()
    window_after = driver.window_handles[0]
    driver.switch_to.window(window_after)
    while True:
        title = driver.title
        if userid not in title:
            driver.switch_to.window(window_after)
        else:
            break
    WebDriverWait(driver, 120).until(
        EC.visibility_of_element_located((By.ID, "preferences")))
    driver.maximize_window()
except Exception as e:
    raise Exception(
        "Error: Unable to login in Agile web service\n\nErr: " + str(e))
for l in range(2,n):
    try:
        item=wks.cell(row=l, column=1).value
        rdoc=wks.cell(row=l,column=2).value
        adoc=wks.cell(row=l,column=3).value
        while True:
            g_ser = driver.find_element_by_id(
                "QUICKSEARCH_STRING").get_attribute("value")
            if item == g_ser:
                break
            else:
                driver.find_element_by_id("QUICKSEARCH_STRING").clear()
                driver.find_element_by_id(
                    "QUICKSEARCH_STRING").send_keys(item)
        driver.find_element_by_id(
            "QUICKSEARCH_STRING").send_keys(Keys.RETURN)
        time.sleep(1)
        while True:
                cursor = driver.find_element_by_tag_name(
                    "body").value_of_css_property("cursor")
                if cursor != "wait":
                    break
        flag=False
        while True:
            try:
                if flag !=True:
                    if driver.find_element_by_class_name('search_header').text=='Search Criteria':
                        n1=len(driver.find_elements_by_class_name("treegrid")[0].find_elements_by_tag_name("table")[0].find_elements_by_class_name("GMBodyMid")[0].find_elements_by_class_name("GMDataRow"))
                        for ele in range(n1):
                            a=driver.find_elements_by_class_name("treegrid")[0].find_elements_by_tag_name("table")[0].find_elements_by_class_name("GMBodyMid")[0].find_elements_by_class_name("GMDataRow")[ele].find_elements_by_tag_name('td')[3].text
                            if str(a).strip()==item.strip():
                                driver.find_elements_by_class_name("treegrid")[0].find_elements_by_tag_name("table")[0].find_elements_by_class_name("GMBodyMid")[0].find_elements_by_class_name("GMDataRow")[ele].find_elements_by_tag_name('td')[3].find_elements_by_tag_name('a')[0].click()
                                flag=True
                                break
                else:
                    n_ele=driver.find_element_by_xpath("//*[@class='column_one layout']").text
                    b=n_ele.split("\n")
                    if b[0]==item:
                        break
            except:
                try:
                    n_ele=driver.find_element_by_xpath("//*[@class='column_one layout']").text
                    b=n_ele.split("\n")
                    if b[0]==item:
                        break
                except:
                    pass
        while True:
                cursor = driver.find_element_by_tag_name(
                    "body").value_of_css_property("cursor")
                if cursor != "wait":
                    break
        while True:
            try:
                n_ele=driver.find_element_by_xpath("//*[@class='column_one layout']").text
                b=n_ele.split("\n")
                if b[0]==item:
                    break
            except:
                pass
        try:
            driver.find_element_by_link_text("Attachments").click()
            WebDriverWait(driver, 120).until(EC.visibility_of_element_located(
                            (By.ID, "view_controls_tabletitle_5")))
            if rdoc != None:
                while True:
                    try:
                        len_rm=len(driver.find_elements_by_class_name('GMBodyMid')[0].find_elements_by_tag_name('tr'))
                        for i in range(len_rm):
                            try:
                                att=driver.find_elements_by_class_name('GMBodyMid')[0].find_elements_by_tag_name('tr')[i].find_elements_by_tag_name('td')[3].text
                                if att.strip()==rdoc:
                                    driver.find_elements_by_class_name('GMBodyLeft')[0].find_elements_by_tag_name('tr')[i].click()
                                    wks.cell(row=l, column=4).value="Remove Pass"
                                    inp.save(loc)
                                    break
                                if i==len_rm:
                                    raise Exception(
                                        "Error: Unable to find the given doc")
                            except:
                                pass
                        break
                    except:
                        pass
                while True:
                    cursor = driver.find_element_by_tag_name(
                        "body").value_of_css_property("cursor")
                    if cursor != "wait":
                        break
                old_count=driver.find_element_by_id('totalCount_ATTACHMENTS_FILELIST').text
                time.sleep(2)
                driver.find_element_by_id('MSG_Remove_5span').click()
                while True:
                    cursor = driver.find_element_by_tag_name(
                        "body").value_of_css_property("cursor")
                    if cursor != "wait":
                        break
                new_count=driver.find_element_by_id('totalCount_ATTACHMENTS_FILELIST').text
                flag1=0
                if old_count!=0:
                    while True:
                        if int(new_count)==int(old_count)-1:
                            break
                        else:
                            new_count=driver.find_element_by_id('totalCount_ATTACHMENTS_FILELIST').text
                        flag1=flag1+1
                        if flag1==10:
                            wks.cell(row=l, column=4).value="Remove Failed"
                            inp.save(loc)
                            raise Exception("Not found")
                        time.sleep(2)
            driver.find_element_by_id('MSG_AddAttachment_5span').click()
            WebDriverWait(driver, 120).until(EC.visibility_of_element_located(
                            (By.ID, "browserFiles")))
            old_count=driver.find_element_by_id('totalCount_ATTACHMENTS_FILELIST').text
            driver.find_element_by_id('browserFiles').click()
            r_file=path + "\\" + adoc
            time.sleep(2)
            while True:
                if win32gui.FindWindow('#32770','Open') != 0:
                    break
            Open_window(r_file)
            time.sleep(2)
            driver.find_element_by_id('uploadFilesUMspan').click()
            time.sleep(10)
            new_count=driver.find_element_by_id('totalCount_ATTACHMENTS_FILELIST').text
            flag1=0
            while True:
                if int(new_count)==int(old_count)+1:
                    wks.cell(row=l, column=5).value="Add Pass"
                    inp.save(loc)
                    break
                else:
                    new_count=driver.find_element_by_id('totalCount_ATTACHMENTS_FILELIST').text
                flag1=flag1+1
                if flag1==10:
                    wks.cell(row=l, column=5).value="Add Failed"
                    inp.save(loc)
                    raise Exception("Not found")
                time.sleep(5)
            driver.find_element_by_id('lfuploadpalette_window_close').click()
        except:
            try:
                driver.find_element_by_id('lfuploadpalette_window_close').click()
            except:
                continue
        #driver.find_element_by_id('lfuploadpalette_window_close').click()
    except Exception as e:
        pass
driver.quit()
messagebox.showinfo("Completed","Process Completed")






    


